from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import configparser

class TextModeler:
    def __new__(cls):
        raise TypeError("TextModeler is a static class used for holding functions for string editing. It cannot be instanced")
    
    @staticmethod
    def cleanSpaces(string):
        return " ".join(string.split())
    
    # @staticmethod
    # @staticmethod
    # def readTime(string):
        
    #     date = str(string.split(" ")[0])
    #     year = date.split("-")[0]
    #     month = date.split("-")[1]
    #     day = date.split("-")[2]

    #     clock = string.split(" ")[1]
    #     hour = time.split(":")[0]
    #     minute = time.split(":")[1]
    #     second = time.split(":")[2]

    #     formatTime = time.struct_time(year, month, day, hour, minute, second)

    #     return time.mktime(formatTime)

class Person:
    def __init__(self, name, email, schoolclass, time, preferences):
        self.mName = name
        self.mEmail = email
        self.mSchoolClass = schoolclass
        self.mTime = time
        self.mPreferences = preferences
    def __str__(self):
        return " ".join([self.mName, self.mEmail, self.mSchoolClass, str(self.mTime), str(self.mPreferences)])

class Sorter:
    def __init__(self):
        self.mTest = "AhoJ svete"
        self.mConfigInfo = configparser.ConfigParser()
        self.mConfigInfo.read("config.ini")
        
        if not self.testConfigFile(self.mConfigInfo):
            raise ImportError("Invalid config.ini file.")

        self.mExReader = ExcelReader(self.mConfigInfo)

    def testConfigFile(self, config):
        if "data" not in config:
            return False
        if "excel header" not in config:
            return False
        return True

class ExcelReader:
    header = {}
    def __init__(self, configinfo):
        # configfileread
        self.mHeaderIsRead = False
        self.mConfig = configinfo
        self.mConfigHeader = self.mConfig["excel header"]
        # excel file read 
        self.mExcelSource = load_workbook(self.mConfig["data"]["filename"])
        self.mExcelWorkSheet = self.mExcelSource.active
        # reading header definitions from config
        self.mHeaderDef = {}
        self.mHeaderDefReverse = {}
        for headerKey in self.mConfigHeader:
            self.mHeaderDef.update({headerKey : self.mConfigHeader.get(headerKey)})
            self.mHeaderDefReverse.update({self.mConfigHeader.get(headerKey) : headerKey})
        self.mHeaderCoord = {}

        self.readHeader()

    def readHeader(self):
        row = 1
        col = 0
        while True:
            col += 1
            coord = get_column_letter(col) + str(row)
            cellvalue = self.mExcelWorkSheet[coord].value
            
            # Stop at 'None' value - empty cell
            if not bool(cellvalue):
                break
            
            # print(cellvalue)
            if cellvalue in self.mHeaderDefReverse:
                print("LOCATED:", cellvalue, "as", self.mHeaderDefReverse[cellvalue])
                self.mHeaderCoord.update( {self.mHeaderDefReverse[cellvalue] : col} )
        print(self.mHeaderCoord)
                
            

        
        







# pers1 = Person("Petr", "pavpetr@fit.cvut.cz", "8.a", datetime.datetime(2022, 6, 23), ["Anna", "Zuzka", "Alzbeta"])

# print(pers1.mPreferences)
# print(pers1)

# excelsource = load_workbook("dialog_data.xlsx")
# ws = excelsource.active
# print(ws["C3"].value)
# print(type(ws["C3"].value))

app = Sorter()