from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import configparser

import Person
from TextModeler import TextModeler

class ExcelReader:
    header = {}
    def __init__(self, configinfo):
        # configfileread
        self.mHeaderIsRead = False
        self.mConfig = configinfo
        # excel file read 
        self.mExcelSource = load_workbook(self.mConfig["data"]["filename"])
        self.mExcelWorkSheet = self.mExcelSource.active

        # for person reading - acts like static value, starts at 1 to skip header
        self.mPersonReadCount = 1

        # read headers and get important coordiantes
        self.readPersonHeaderCord()
        self.readSlotHeaderCord()

    def readPersonHeaderCord(self):
        """Reads whole header and locates column coordinates for personal header columns"""
        personConfig = self.mConfig["person"]
        personHeaderDef = {}
        self.mPersonHeaderCoord = {}
        for headerKey in personConfig:
            personHeaderDef.update({personConfig.get(headerKey) : headerKey})

        row, col = 1, 0
        while True:
            col += 1
            coord = get_column_letter(col) + str(row)
            cellvalue = self.mExcelWorkSheet[coord].value
            
            # Stop at 'None' value - empty cell
            if not bool(cellvalue):
                break
            
            if cellvalue in personHeaderDef:
                print("LOCATED:", cellvalue, "as", personHeaderDef[cellvalue])
                self.mPersonHeaderCoord.update( {personHeaderDef[cellvalue] : col} )
        # warning print for missing parameters
        for headerKey in personConfig:
            if headerKey not in self.mPersonHeaderCoord:
                print("Header:", headerKey, "not located!")
    
    def readSlotHeaderCord(self):
        """Reads whole header and locates column coordinates for slot header columns"""
        slotConfig = self.mConfig["slot"]
        slotHeaderDef = {}
        self.mSlotHeaderCoord = {}
        for headerKey in slotConfig:
            slotHeaderDef.update({slotConfig.get(headerKey) : headerKey})

        row, col = 1, 0
        while True:
            col += 1
            coord = get_column_letter(col) + str(row)
            cellvalue = self.mExcelWorkSheet[coord].value
            
            # Stop at 'None' value - empty cell
            if not bool(cellvalue):
                break
            
            if cellvalue in slotHeaderDef:
                print("LOCATED:", cellvalue, "as", slotHeaderDef[cellvalue])
                self.mSlotHeaderCoord.update( {slotHeaderDef[cellvalue] : col} )
        # warning print for missing parameters
        for headerKey in slotConfig:
            if headerKey not in self.mSlotHeaderCoord:
                print("Header:", headerKey, "not located!")

    def readPerson(self, personNumber=0):
        personalInfo = {}
        preferences = {}
        if personNumber == 0:
            self.mPersonReadCount += 1
            row = self.mPersonReadCount
        else:
            # +1 to skip header -> personNumber 1 == first person in table
            row = personNumber + 1 

        for personInformation in self.mPersonHeaderCoord:
            col = self.mPersonHeaderCoord[personInformation]
            coord = get_column_letter(col) + str(row)
            cellvalue = self.mExcelWorkSheet[coord].value
            cellvalue = TextModeler.cleanSpaces(cellvalue)
            
            personalInfo.update( {personInformation : cellvalue} )

        for slotInformation in self.mSlotHeaderCoord:
            col = self.mSlotHeaderCoord[slotInformation]
            coord = get_column_letter(col) + str(row)
            cellvalue = self.mExcelWorkSheet[coord].value
            cellvalue = TextModeler.cleanSpaces(cellvalue)
            
            preferences.update( {slotInformation : cellvalue} )


        if bool(personalInfo["name"]) is False:
            return False
        
        return Person.Person(personalInfo, preferences) 