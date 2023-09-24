from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import configparser

import Person
from TextModeler import TextModeler

class ExcelPrinter:
    def __init__(self, config):
        self.mConfig = config

        # reading origin file cause we want to add new sheets to it
        self.mExcelFile = load_workbook(self.mConfig["data"]["filename"])
    
    def outputActivitySheets(self, activities):
        for activity in activities:
            wsheet = self.mExcelFile.create_sheet(activity.mName)
            wsheet.title = activity.mName
            wsheet["A1"].value = activity.mName
            wsheet["B1"].value = "Applicants count:"
            wsheet["C1"].value = len(activity.mApplicants)
            wsheet["D1"].value = "Asigned capacity:"
            wsheet["E1"].value = activity.mCapacity

            headerInfo = activity.mApplicants[0].outputHeaderInfo()
            # prints a header for person information on row 2
            row = 2
            for i in range( len(headerInfo) ):
                cell = wsheet.cell(row = row, column=i + 1)
                cell.value = headerInfo[i]
            
            for person in activity.mApplicants:
                # print(person)
                row += 1
                column = 1
                for information in person.outputPersonalInfo():
                    cell = wsheet.cell(row = row, column = column)
                    cell.value = information
                    column += 1
                # plus 1 as slots are one-indexed
                for slotNumber in range(1, len(person.mPreferences) + 1):
                    cell = wsheet.cell(row = row, column = column)
                    cell.value = person.mActivities.get(slotNumber, "NOT ASIGNED")
                    column += 1
        return
    
    def outputPeopleSheet(self, people):
        wsheet = self.mExcelFile.create_sheet(self.mConfig["output"]["peoplesheetname"])
        wsheet.title = self.mConfig["output"]["peoplesheetname"]
        wsheet["A1"].value = self.mConfig["output"]["peoplesheetname"]
        wsheet["B1"].value = "Applicants count:"
        wsheet["C1"].value = len(people)

        headerInfo = people[0].outputHeaderInfo()
        # prints a header for person information on row 2
        row = 2
        for i in range( len(headerInfo) ):
            cell = wsheet.cell(row = row, column=i + 1)
            cell.value = headerInfo[i]
        
        for person in people:
            # print(person)
            row += 1
            column = 1
            for information in person.outputPersonalInfo():
                cell = wsheet.cell(row = row, column = column)
                cell.value = information
                column += 1
            # plus 1 as slots are one-indexed
            for slotNumber in range(1, len(person.mPreferences) + 1):
                cell = wsheet.cell(row = row, column = column)
                cell.value = person.mActivities.get(slotNumber, "NOT ASIGNED")
                column += 1
        return

    def saveFile(self):
        oldname = self.mConfig["data"]["filename"]
        oldname = oldname.split(".")
        filename = oldname[0] + "_SORTED.xlsx"
        
        self.mExcelFile.save(filename)
        return