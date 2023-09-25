from openpyxl import load_workbook

import configparser as cp

from Person import Person
from TextModeler import TextModeler


class ExcelPrinter:
    """Helper class used for outputing a new excel file with generated sheets

    Attributes:
        mConfig: configfile, required for filename reading
        mExcelFile: openpyxl excelfile
    """

    mConfig = cp.ConfigParser()
    mExcelFile = None

    def __init__(self, config: cp.ConfigParser):
        """constructs a new printer, uses filename from config to preserver original data in output

        Loads the original file

        Args:
            config (configParser): whole config file is expected
        """
        self.mConfig = config

        # reading origin file cause we want to add new sheets to it
        self.mExcelFile = load_workbook(self.mConfig["data"]["filename"])

    def outputActivitySheets(self, activities: list) -> None:
        """Outputs new sheets for every activity in given list

        Every sheet contains information about the activity and a list of attendace with all personal info
        and even asigned activites to the person.

        Args:
            activities (list): list of Activities

        Returns:
            None : Everything is written to mExcelFile
        """
        for activity in activities:
            wsheet = self.mExcelFile.create_sheet(activity.mName)
            wsheet.title = activity.mName
            wsheet["A1"].value = activity.mName
            wsheet["B1"].value = "Applicants count:"
            wsheet["C1"].value = len(activity.mApplicants)
            wsheet["D1"].value = "Asigned capacity:"
            wsheet["E1"].value = activity.mCapacity

            headerInfo = activity.mApplicants[0].outputHeaderInfo()
            # prints a header for person information on row 2
            row = 2
            for i in range(len(headerInfo)):
                cell = wsheet.cell(row=row, column=i + 1)
                cell.value = headerInfo[i]

            for person in activity.mApplicants:
                # print(person)
                row += 1
                column = 1
                for information in person.outputPersonalInfo():
                    cell = wsheet.cell(row=row, column=column)
                    cell.value = information
                    column += 1
                for slotinfo in person.outputAsignedSlots():
                    cell = wsheet.cell(row=row, column=column)
                    cell.value = slotinfo
                    column += 1
        return

    def outputPeopleSheet(self, people) -> None:
        """Outputs new sheet with all the people and their asigned activities.

        Every sheet contains all personal info
        and even asigned activites to the person.

        Args:
            people (list): list of people

        Returns:
            None : Everything is written to mExcelFile
        """
        wsheet = self.mExcelFile.create_sheet(self.mConfig["output"]["peoplesheetname"])
        wsheet.title = self.mConfig["output"]["peoplesheetname"]
        wsheet["A1"].value = self.mConfig["output"]["peoplesheetname"]
        wsheet["B1"].value = "Applicants count:"
        wsheet["C1"].value = len(people)

        headerInfo = people[0].outputHeaderInfo()
        # prints a header for person information on row 2
        row = 2
        for i in range(len(headerInfo)):
            cell = wsheet.cell(row=row, column=i + 1)
            cell.value = headerInfo[i]

        for person in people:
            # print(person)
            row += 1
            column = 1
            for information in person.outputPersonalInfo():
                cell = wsheet.cell(row=row, column=column)
                cell.value = information
                column += 1
            for slotinfo in person.outputAsignedSlots():
                cell = wsheet.cell(row=row, column=column)
                cell.value = slotinfo
                column += 1
        return

    def saveFile(self) -> None:
        """Saves the file in syntax "ORIGINALNAME_SORTED.xlsx"""
        oldname = self.mConfig["data"]["filename"]
        oldname = oldname.split(".")
        filename = oldname[0] + "_SORTED.xlsx"

        self.mExcelFile.save(filename)
        return
